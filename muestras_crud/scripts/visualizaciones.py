import os
import matplotlib.pyplot as plt
import pandas as pd
import numpy as np
from scipy.ndimage import gaussian_filter1d
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import os
def draw_histogram(df, columna, archivo_planilla, nombre_hoja, bins=15):
    """
    Genera un histograma de la columna indicada y lo inserta en la hoja Excel especificada.
    """



    # Asegurar que la columna sea numérica
    data = pd.to_numeric(df[columna], errors='coerce').dropna()

    # Crear histograma
    plt.figure(figsize=(10, 6))
    plt.hist(data, bins=bins, color="skyblue", edgecolor="black")
    plt.title(f"Histograma de {columna}")
    plt.xlabel(columna)
    plt.ylabel("Frecuencia")
    plt.grid(axis="y", linestyle="--", alpha=0.7)
    plt.tight_layout()

    # Guardar imagen temporal
    temp_image_path = f"hist_{columna}.png"
    plt.savefig(temp_image_path, dpi=300)
    plt.close()

    # Insertar en Excel
    if os.path.exists(archivo_planilla):
        workbook = load_workbook(archivo_planilla)
        if nombre_hoja not in workbook.sheetnames:
            workbook.create_sheet(nombre_hoja)
        hoja = workbook[nombre_hoja]

        img = Image(temp_image_path)
        img.anchor = "A20"  # un poco más abajo para no pisar otras gráficas
        hoja.add_image(img)

        workbook.save(archivo_planilla)
        workbook.close()
        os.remove(temp_image_path)
        print(f"Histograma '{columna}' insertado en la hoja '{nombre_hoja}'.")
    else:
        print(f"No existe el archivo '{archivo_planilla}'. No se insertó histograma.")
# === Utilidad para crear carpetas ===
def ensure_dir(path):
    if not os.path.exists(path):
        os.makedirs(path)

def draw_DFT(df):
    try:
        x = pd.to_numeric(df.iloc[:, 0], errors='coerce')
        y = pd.to_numeric(df.iloc[:, 3], errors='coerce')
        mask = x.notna() & y.notna()
        x, y = x[mask].values, y[mask].values
        y_scaled = y * 100000

        plt.figure(figsize=(10, 6))
        widths = np.diff(x, append=x[-1] + (x[-1]-x[-2]))
        plt.bar(x, y_scaled, width=widths, align='edge', color='#228B22', edgecolor='black', linewidth=0.3)
        plt.xlabel('Half pore width, Å')
        plt.ylabel(r'Pore volume $\times 10^{-5}$, cc $\AA^{-1}$ g$^{-1}$')
        plt.title('PAH (DFT Analysis)', loc='left', fontweight='bold')
        plt.xlim(0, 180)
        plt.tight_layout()

        output_name = "grafico_dft_final.png"
        plt.savefig(output_name, dpi=300)
        plt.close()
        print(f"DFT guardado como: {output_name}")
        return output_name   # <-- RETORNAR la ruta
    except Exception as e:
        print(f"Error en DFT: {e}")
        return None

# Dentro de visualizaciones.py, cambia el inicio de draw_HK:
def draw_HK(df):
    try:
        x = pd.to_numeric(df.iloc[:, 0], errors='coerce')
        y = pd.to_numeric(df.iloc[:, 1], errors='coerce')
        mask = x.notna() & y.notna()
        x, y = x[mask].values, y[mask].values

        plt.figure(figsize=(10, 6))
        plt.step(x, y, where='post', color='forestgreen', linewidth=1.5)
        plt.fill_between(x, y, step="post", color='forestgreen', alpha=0.2)
        plt.xlabel('Half pore width, Å')
        plt.ylabel(r'dV, cc $\AA^{-1}$ g$^{-1}$')
        plt.title('Distribución de Tamaño de Poros (HK)', fontweight='bold')
        plt.grid(True, linestyle='--', alpha=0.3)
        plt.tight_layout()

        output_name = "grafico_hk_final.png"
        plt.savefig(output_name, dpi=300)
        plt.close()
        print(f"HK guardado como: {output_name}")
        return output_name   # <-- RETORNAR la ruta
    except Exception as e:
        print(f"Error en HK: {e}")
        return None
# === BET ===
def draw_BET(df, output_dir=None):
    output_dir = output_dir or os.path.join(os.getcwd(), "graficos_temp")
    ensure_dir(output_dir)
    ruta_img = os.path.join(output_dir, "bet.png")

    try:
        col_x = next((c for c in df.columns if "p" in c.lower()), df.columns[0])
        col_y = next((c for c in df.columns if "volume" in c.lower() or "stp" in c.lower()), df.columns[1])

        x = pd.to_numeric(df[col_x], errors="coerce")
        y = pd.to_numeric(df[col_y], errors="coerce")
        mask = x.notna() & y.notna()
        x, y = x[mask], y[mask]

        if len(x) == 0:
            print("[warn] draw_BET: No hay datos válidos")
            return None

        plt.figure(figsize=(6, 4))
        plt.plot(x, y, marker="o", color="purple", linewidth=1.2)
        plt.xlabel(col_x)
        plt.ylabel(col_y)
        plt.title("Isoterma BET")
        plt.grid(True, linestyle="--", alpha=0.6)
        plt.tight_layout()
        plt.savefig(ruta_img, dpi=150)
        plt.close()
        return ruta_img
    except Exception as e:
        print(f"[error] draw_BET: {e}")
        return None


def draw_comparison_bar_chart(df_ads, df_des, output_dir=None):
    output_dir = output_dir or os.getcwd()
    ensure_dir(output_dir)
    ruta_img = os.path.join(output_dir, "grafico_bjh_final.png")

    try:
        def get_clean_xy(df):
            # Limpieza profunda: convertir a número y quitar filas vacías
            df_numeric = df.apply(pd.to_numeric, errors='coerce').dropna(subset=[df.columns[0], df.columns[1]])
            x = df_numeric.iloc[:, 0].values / 10  # radio en nm
            y = df_numeric.iloc[:, 1].values
            # Filtro para evitar errores de log10
            mask = (x > 0) & (y > 1e-15) 
            return x[mask], np.log10(y[mask])

        x_ads, y_ads = get_clean_xy(df_ads)
        x_des, y_des = get_clean_xy(df_des)

        plt.figure(figsize=(6, 4))

        # --- DIBUJAR ADSORCIÓN (ROJO) ---
        if len(x_ads) > 0:
            # Línea escalonada
            plt.step(x_ads, y_ads, where='post', color="red", label="adsorption", linewidth=1.0)
            # Líneas verticales para efecto de barra
            plt.vlines(x_ads, -7, y_ads, colors='red', linewidth=0.6, alpha=0.8)

        # --- DIBUJAR DESORCIÓN (AZUL) ---
        if len(x_des) > 0:
            # Línea escalonada
            plt.step(x_des, y_des, where='post', color="blue", label="desorption", linewidth=1.0)
            # Líneas verticales para efecto de barra
            plt.vlines(x_des, -7, y_des, colors='blue', linewidth=0.6, alpha=0.8)

        # Configuración de ejes igual al reporte
        plt.xscale('log')
        plt.xlabel("Pore radius, nm", fontsize=10)
        plt.ylabel(r"log dV, cc g$^{-1}$", fontsize=10)
        plt.title("PAH", loc='left', fontsize=12, fontweight='bold')
        
        plt.xlim(1.5, 200)
        plt.ylim(-7, -3) 
        
        plt.legend(frameon=False, loc='upper right')
        plt.grid(False)
        plt.tight_layout()

        plt.savefig(ruta_img, dpi=300)
        plt.close()
        
        print(f"Gráfico generado con Adsorción y Desorción: {ruta_img}")
        return ruta_img

    except Exception as e:
        print(f"[error] draw_comparison_bar_chart: {e}")
        return None