# Graficos a partir de los dataframes
import openpyxl
import os
import csv
import traceback
from datetime import datetime
from pywinauto import Application, findwindows
import time
import threading
from novawinmng import (manejar_novawin, leer_csv_y_crear_dataframe, 
                        agregar_csv_a_plantilla_excel, guardar_dataframe_en_ini, 
                        generar_nombre_unico, agregar_dataframe_a_excel_sin_borrar, 
                        agregar_dataframe_a_nueva_hoja)
from pywinauto.keyboard import send_keys
from openpyxl import Workbook
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

def draw_histogram(df, columna, archivo_planilla, nombre_hoja, bins=15):
    """
    Genera un histograma de la columna indicada y lo inserta en la hoja Excel especificada.
    """
    data = pd.to_numeric(df[columna], errors='coerce').dropna()

    plt.figure(figsize=(10, 6))
    plt.hist(data, bins=bins, color="skyblue", edgecolor="black")
    plt.title(f"Histograma de {columna}")
    plt.xlabel(columna)
    plt.ylabel("Frecuencia")
    plt.grid(axis="y", linestyle="--", alpha=0.7)
    plt.tight_layout()

    temp_image_path = f"hist_{columna}.png"
    plt.savefig(temp_image_path, dpi=300)
    plt.close()

    if os.path.exists(archivo_planilla):
        workbook = load_workbook(archivo_planilla)
        if nombre_hoja not in workbook.sheetnames:
            workbook.create_sheet(nombre_hoja)
        hoja = workbook[nombre_hoja]

        img = Image(temp_image_path)
        img.anchor = "A20"  
        hoja.add_image(img)

        workbook.save(archivo_planilla)
        workbook.close()
        os.remove(temp_image_path)
        print(f"Histograma '{columna}' insertado en la hoja '{nombre_hoja}'.")
    else:
        print(f"No existe el archivo '{archivo_planilla}'.")

def draw_comparison_bar_chart(bjhd, bjha, archivo_planilla):
    radius_bjhd = bjhd['Radius']
    dV_logr_bjhd = bjhd['dV(logr)']
    radius_bjha = bjha['Radius']
    dV_logr_bjha = bjha['dV(logr)']
    
    min_length = min(len(radius_bjhd), len(radius_bjha))
    radius_bjhd = radius_bjhd[:min_length]
    dV_logr_bjhd = dV_logr_bjhd[:min_length]
    radius_bjha = radius_bjha[:min_length]
    dV_logr_bjha = dV_logr_bjha[:min_length]

    plt.figure(figsize=(12, 6))
    bar_width = 0.5
    index = range(len(radius_bjhd))
    
    plt.bar(index, dV_logr_bjhd, bar_width, label='Desorbcion', edgecolor='blue', fill=False, linewidth=1.5)
    plt.bar([i + bar_width for i in index], dV_logr_bjha, bar_width, label='Absorpcion', edgecolor='green', fill=False, linewidth=1.5)

    plt.xlabel('Radius (Å)')
    plt.ylabel('dV(logr) (cc/Å/g)')
    plt.title('BJH Absorpcion y Desorbcion')
    plt.xticks([i + bar_width / 2 for i in index], radius_bjhd, rotation=90)
    plt.legend()
    plt.grid(axis='y', linestyle='--', alpha=0.6)
    plt.tight_layout()
    
    temp_image_path = "grafico_bjh_comp.png"
    plt.savefig(temp_image_path, dpi=300)
    plt.close()

    if os.path.exists(archivo_planilla):
        workbook = load_workbook(archivo_planilla)
        if "BJH" not in workbook.sheetnames:
            workbook.create_sheet("BJH")
        hoja = workbook["BJH"]
        img = Image(temp_image_path)
        img.anchor = 'G1'
        hoja.add_image(img)
        workbook.save(archivo_planilla)
        workbook.close()
        os.remove(temp_image_path)
        print(f"Gráfico BJH insertado en '{archivo_planilla}'.")

def draw_DFT(df, archivo_planilla):
    half_pore_width = pd.to_numeric(df['Half pore width'], errors='coerce')
    dVr = df['dV(r)']
    
    valid_indices = ~half_pore_width.isna()
    dVr = dVr[valid_indices]
    half_pore_width = half_pore_width[valid_indices]

    plt.figure(figsize=(12, 6))
    plt.bar(half_pore_width, dVr, width=0.8, color='green', alpha=0.7, label='dV(r)')
    plt.xlabel('Half pore width (Å)')
    plt.ylabel('dV(r) (cc/Å/g)')
    plt.title('DFT Analysis: Half Pore Width vs dV(r)')
    plt.grid(axis='y', linestyle='--', alpha=0.6)
    plt.tight_layout()

    temp_image_path = "grafico_dft.png"
    plt.savefig(temp_image_path, dpi=300)
    plt.close()

    if os.path.exists(archivo_planilla):
        workbook = load_workbook(archivo_planilla)
        if "DFT" not in workbook.sheetnames:
            workbook.create_sheet("DFT")
        hoja = workbook["DFT"]
        img = Image(temp_image_path)
        img.anchor = 'G1'
        hoja.add_image(img)
        workbook.save(archivo_planilla)
        workbook.close()
        os.remove(temp_image_path)
        draw_histogram(df, "Half pore width", archivo_planilla, "DFT")

def draw_HK(df, archivo_planilla):
    dVr = df['dV()']
    half_pore_width = pd.to_numeric(df['Half pore width'], errors='coerce')

    valid_indices = ~half_pore_width.isna()
    dVr = dVr[valid_indices]
    half_pore_width = half_pore_width[valid_indices]

    plt.figure(figsize=(14, 6))
    plt.bar(half_pore_width, dVr, width=0.8, color='green', alpha=0.7, label='dV(r)')
    plt.xlabel('Half pore width , A')
    plt.ylabel('dV cc A g')
    plt.title('Gráfico HK')
    plt.xticks(half_pore_width, rotation=90)
    plt.tight_layout()

    temp_image_path = "grafico_hk.png"
    plt.savefig(temp_image_path, dpi=300)
    plt.close()

    if os.path.exists(archivo_planilla):
        workbook = load_workbook(archivo_planilla)
        if "HK" not in workbook.sheetnames:
            workbook.create_sheet("HK")
        hoja = workbook["HK"]
        img = Image(temp_image_path)
        img.anchor = 'A1'
        hoja.add_image(img)
        workbook.save(archivo_planilla)
        workbook.close()
        os.remove(temp_image_path)
        draw_histogram(df, "Half pore width", archivo_planilla, "HK")

def graphs_main(archivo_planilla, hoja='NKA'):
    print("Inicio de graphs_main")
    
    # Consolidación de NKA y FFHA
    df_nka = pd.read_excel(archivo_planilla, sheet_name=hoja)
    df_ffha = pd.read_excel(archivo_planilla, sheet_name=hoja) # O 'FFHA' si corresponde
    nuevo_df = pd.concat([df_nka.tail(2).reset_index(drop=True), 
                         df_ffha.tail(2).reset_index(drop=True)], 
                        keys=['NKA', 'FFHA'], names=['Sheet', 'Index'])
    
    try:
        book = openpyxl.load_workbook(archivo_planilla)
        if 'NKAFFHA_valores' in book.sheetnames:
            del book['NKAFFHA_valores']
        book.save(archivo_planilla)
    except FileNotFoundError:
        pass

    with pd.ExcelWriter(archivo_planilla, engine='openpyxl', mode='a') as writer:
        nuevo_df.to_excel(writer, sheet_name='NKAFFHA_valores', index=False)

    # Procesar HK
    df_hk = pd.read_excel(archivo_planilla, sheet_name='HK')
    draw_HK(df_hk, archivo_planilla)

    # Procesar DFT
    df_dft = pd.read_excel(archivo_planilla, sheet_name='DFT')
    draw_DFT(df_dft, archivo_planilla)

    # Procesar BJH
    df_bjha = pd.read_excel(archivo_planilla, sheet_name='BJHA')
    df_bjhd = pd.read_excel(archivo_planilla, sheet_name='BJHD')
    draw_comparison_bar_chart(df_bjhd, df_bjha, archivo_planilla)

    print("Proceso de graficación finalizado. No se ejecutará Novawin.")