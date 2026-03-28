import logging
import sys
from pathlib import Path
import pandas as pd
from tests_bet import _format_valor
import numpy as np
import matplotlib.pyplot as plt
from scipy.stats import linregress
from modelo_fractal import fmain
from openpyxl import load_workbook

def guardar_resultados_completos(ruta_excel, resultados_dict):
    wb = load_workbook(ruta_excel)
    nombre_hoja = "RESULTADO_CALCULOS"
    
    if nombre_hoja in wb.sheetnames:
        ws = wb[nombre_hoja]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(nombre_hoja)

    ws.append(["Parámetro", "Valor"])

    def f(v):
        if v is None or pd.isna(v):
            return "nd"
        return _format_valor(v)

    # Fractales y BET
    ws.append(["FHH", f(resultados_dict.get("FHH"))])
    ws.append(["NK", f(resultados_dict.get("NK"))])
    ws.append(["SingleBET", f(resultados_dict.get("SingleBET"))])
    ws.append(["BET_SD", f(resultados_dict.get("BET_SD"))])

    # Volúmenes por rango
    vol_por_rango = resultados_dict.get("Volumenes_por_rango", {})
    ws.append(["Volumen de microporos (1.6-9.7 Å), cc g-1", f(vol_por_rango.get("Microporos_1.6-9.7Å"))])
    ws.append(["Volumen de mesoporos (7.7-180 Å), cc g-1", f(vol_por_rango.get("Mesoporos_7.7-180Å"))])
    ws.append(["Volumen de mesoporos (15-1767 Å), cc g-1", f(vol_por_rango.get("Mesoporos_15-1767Å"))])

    wb.save(ruta_excel)
    log("RESULTADO_CALCULOS escrito", "info")

def normalizar_dft_poros(volumenes_por_rango):
    dft_poros = {}
    for pore, valor in volumenes_por_rango.items():
        dft_poros[pore] = {"Total": float(valor)}
    return dft_poros
def guardar_resultados_poros(dft_poros, archivo_salida, valor="Total"):
    """
    Guarda DFT_Poros en Excel en una hoja 'RESULTADO_POROS' sin borrar otras hojas.
    """
    # Convertir a DataFrame
    filas = []
    for pore, datos in dft_poros.items():
        v = datos.get(valor, "nd")
        if v == 0.0 or v == "nd":
            v = "nd"
        filas.append({"Pore Type": pore, "Value": v})

    df_result = pd.DataFrame(filas)

    # Abrir workbook existente
    wb = load_workbook(archivo_salida)

    # Si la hoja existe, borrarla
    sheet_name = "RESULTADO_POROS"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    # Crear hoja nueva
    ws = wb.create_sheet(sheet_name)

    # Escribir DataFrame en la hoja
    for r_idx, row in enumerate(df_result.itertuples(index=False), start=2):
        ws.cell(row=r_idx, column=1, value=row[0])
        ws.cell(row=r_idx, column=2, value=row[1])

    # Escribir encabezado
    ws.cell(row=1, column=1, value="Pore Type")
    ws.cell(row=1, column=2, value="Value")

    wb.save(archivo_salida)
    log(f"RESULTADO_POROS (DFT) escrito correctamente: {archivo_salida}")
def setup_logging(force=True):
    import logging, sys
    from pathlib import Path

    SCRIPT_DIR = Path(__file__).resolve().parent
    LOGS_DIR = SCRIPT_DIR / "logs"
    LOGS_DIR.mkdir(parents=True, exist_ok=True)
    LOG_FILE = LOGS_DIR / "proceso_novawin.log"

    root = logging.getLogger()

    if force:
        for h in root.handlers[:]:
            root.removeHandler(h)

    root.setLevel(logging.DEBUG)

    formatter = logging.Formatter(
        "%(asctime)s.%(msecs)03d | %(levelname)s | %(name)s | %(message)s",
        "%Y-%m-%d %H:%M:%S"
    )

    fh = logging.FileHandler(LOG_FILE, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(formatter)

    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(formatter)

    root.addHandler(fh)
    root.addHandler(ch)

    print("LOG FILE REAL:", LOG_FILE)

def log(msg, level="info"):
    level = level.lower()
    logger = logging.getLogger()  # ROOT

    if level == "debug":
        logger.debug(msg)
    elif level in ("warn", "warning"):
        logger.warning(msg)
    elif level == "error":
        logger.error(msg)
    elif level == "critical":
        logger.critical(msg)
    else:
        logger.info(msg)

    for h in logger.handlers:
        try:
            h.flush()
        except Exception:
            pass

setup_logging(force=True)
logging.getLogger("PIL").setLevel(logging.WARNING)
log("Logging inicializado (inicio del archivo)")       
        
import django
from pywinauto import Application
from pywinauto.keyboard import send_keys
from openpyxl import load_workbook
from django.core.files import File
import ctypes
import os  
import time      
# ===================================================
# 2. RUTAS Y DJANGO
# ===================================================
current_dir = Path(__file__).resolve()
root_dir = current_dir
while root_dir.name != 'muestras_crud' and str(root_dir) != str(current_dir.root):
    root_dir = root_dir.parent

PROJECT_CONTAINER = root_dir.parent if root_dir.name == 'muestras_crud' else current_dir.parent.parent
sys.path.append(str(PROJECT_CONTAINER))
print("ANTES django.setup ", logging.getLogger().handlers)

    
def agregar_hoja_resultados(ruta_excel_principal, ruta_excel_resultados):
    """
    Carga el archivo Excel de resultados BET y copia la primera hoja 
    (que contiene la tabla final) al archivo Excel principal.
    """
    if not ruta_excel_resultados or not os.path.exists(ruta_excel_resultados):
        log("No se encontró la ruta del archivo de resultados BET. Omitiendo la fusión.", "warn")
        return ruta_excel_principal

    log(f"Iniciando fusión de resultados BET desde '{ruta_excel_resultados}' a '{ruta_excel_principal}'")
    
    try:
        # Cargar el archivo principal
        wb_principal = load_workbook(ruta_excel_principal)
        
        # Cargar el archivo de resultados
        wb_resultados = load_workbook(ruta_excel_resultados)
        
        # Asumimos que la hoja de resultados es la primera (índice 0)
        ws_resultados = wb_resultados.worksheets[0]
        
        # 1. Crear una nueva hoja en el principal
        nombre_hoja_nueva = "RESULTADOS_BET"
        if nombre_hoja_nueva in wb_principal.sheetnames:
            log(f"La hoja '{nombre_hoja_nueva}' ya existe. Eliminándola.", "warn")
            del wb_principal[nombre_hoja_nueva]
            
        ws_nueva = wb_principal.create_sheet(nombre_hoja_nueva, index=len(wb_principal.sheetnames))
        
        # 2. Copiar todas las celdas de la hoja de resultados
        for row in ws_resultados.iter_rows():
            for cell in row:
                ws_nueva[cell.coordinate].value = cell.value
                
        # 3. Guardar el archivo principal modificado
        wb_principal.save(ruta_excel_principal)
        log(f"Hoja '{nombre_hoja_nueva}' agregada exitosamente al archivo principal.", "info")
        
        # Opcional: Eliminar el archivo temporal de resultados
        os.remove(ruta_excel_resultados)
        log(f"Archivo de resultados temporal eliminado: {ruta_excel_resultados}", "debug")
        
        return ruta_excel_principal
        
    except Exception as e:
        log(f"Error al agregar la hoja de resultados BET: {e}", "error")
        # Es CRÍTICO que el proceso continúe si esto falla.
        return ruta_excel_principal
def subir_resultados_a_django(empresa_nombre, encargado, sample_number, ruta_xlsx, ruta_pdf):
    # ... (cuerpo de la función) ...
    empresa, _ = Empresa.objects.get_or_create(
        nombre=empresa_nombre,
        defaults={"encargado": encargado}
    )
    muestra, _ = Muestra.objects.get_or_create(
        empresa=empresa,
        sample_number=sample_number
    )
    inf = Informe.objects.create(muestra=muestra)

    # Carpeta relativa dentro de MEDIA_ROOT (Esto está bien, se maneja por Django)
    xlsx_dest = "informes/xlsx"
    pdf_dest  = "informes/pdf"

    # Guardar Excel
    if os.path.exists(ruta_xlsx):
        with open(ruta_xlsx, "rb") as f:
            inf.archivo_xlsx.save(
                os.path.basename(ruta_xlsx),
                File(f),
                save=True
            )
    else:
        log(f"No se encontró el archivo Excel: {ruta_xlsx}", "error")

    # Guardar PDF
    if os.path.exists(ruta_pdf):
        with open(ruta_pdf, "rb") as f:
            inf.archivo_pdf.save(
                os.path.basename(ruta_pdf),
                File(f),
                save=True
            )
    else:
        log(f"No se encontró el archivo PDF: {ruta_pdf}", "error")

    return inf
# -----------------------------
# Funciones auxiliares
# -----------------------------
def renombrar_y_reordenar_hojas(ruta_excel):
    nombres_correctos = ["INFORME", "RESULTADO_CALCULOS", "BET", "BJHA", "BJHD", "HK", "DFT",
                         "NKA", "FHHA"]
    wb = load_workbook(ruta_excel)
    hojas = wb.sheetnames
    log(f"[debug] hojas detectadas: {hojas}", "debug")

    for i, nombre in enumerate(nombres_correctos):
        if i < len(hojas):
            wb[hojas[i]].title = nombre
        else:
            wb.create_sheet(title=nombre)

    for i, nombre in enumerate(wb.sheetnames[len(nombres_correctos):], start=1):
        wb[nombre].title = f"Extra_{i}"

    wb.save(ruta_excel)
    log(f"[debug] hojas finales: {wb.sheetnames}", "debug")
    return ruta_excel

def cerrar_novawin(app):
    log("Cerrando NovaWin...", "warn")
    try:
        # Verificamos si app es un objeto Application válido
        if app and hasattr(app, 'is_running') and app.is_running():
            app.kill()
            log("NovaWin cerrado con kill().")
    except Exception as e:
        log(f"Error al cerrar: {e}", "error")

def iniciar_novawin(path_novawin):
    for backend in ["win32", "uia"]:
        try:
            log(f"Intentando backend {backend}...")
            app = Application(backend=backend).start(path_novawin)
            time.sleep(5)
            ventana = app.window(title_re=".*NovaWin.*")
            ventana.wait("exists", timeout=10)
            log(f"NovaWin iniciado con {backend}")
            return app, ventana
        except Exception as e:
            log(f"Fallo backend {backend}: {e}", "debug")
    raise RuntimeError("No se pudo iniciar NovaWin")

def abrir_qps(ventana, path_qps):
    log(f"Abriendo QPS: {path_qps}")
    ventana.set_focus()
    send_keys('%fo')
    time.sleep(1)
    send_keys(path_qps + '{ENTER}')
    log("Comando de apertura enviado.")
    


def exportar_reporte(graph_view, carpeta_export, rname):

    log(f"[debug] Exportando desde la vista: title='{graph_view.window_text()}'")

   

    # 1. Exportar a ASCII: Abrir el menú contextual

   

    # Intentar el click derecho y, si falla, usar el teclado (APPSKEY)

    menu_abierto = False

    try:

        # Intento A: Clic derecho directo (el método más limpio si funciona)

        graph_view.right_click_input()

        menu_abierto = True
        time.sleep(1)
        log("[debug] Menú contextual abierto con click derecho.")

       

    except Exception as e:

        # Intento B: Si falla el click, forzar foco y usar la tecla de menú contextual

        log(f"[Advertencia] Fallo al hacer click derecho: {e}. Usando tecla de menú (APPSKEY).", "error")

        try:

            graph_view.set_focus()

            time.sleep(0.5)

            send_keys('{APPSKEY}') # Tecla de Menú contextual (simula Shift+F10)

            time.sleep(1) # Dar tiempo al menú para aparecer

            menu_abierto = True

        except Exception as focus_error:

            log(f"Error fatal: No se pudo establecer el foco para usar APPSKEY: {focus_error}", "error")

            return None # Fallo en la exportación

           

    if menu_abierto:

        # Si el menú contextual está visible, enviar la secuencia de Exportar/ASCII

        send_keys('p') # Export (Verifica si 'P' es el atajo)

        time.sleep(1)

        send_keys('p') # ASCII (Verifica si 'P' es el atajo)

        time.sleep(1)

        send_keys('{ENTER}') # Confirmar la selección de ASCII (si aplica)

        time.sleep(1) # Espera a que el diálogo de guardado aparezca

        graph_view.right_click_input()
        time.sleep(1)
        send_keys('v') # ASCII (Verifica si 'P' es el atajo)
    # 2. Guardar archivo TXT

    # ... (El resto del código de guardado se mantiene igual, ya que es la mejor práctica de pywinauto) ...

    # ... (Se mantiene el manejo de GuardarButton y el fallback con send_keys) ...



    try:

        # Conectarse al nuevo diálogo "Guardar como"

        guardar_como = Application().connect(title_re=".*Guardar como|.*Save As", timeout=5).top_window()

        ruta_completa = os.path.join(carpeta_export, f"{rname}_export.txt")

        guardar_como["Edit"].set_text(ruta_completa)

        guardar_como["GuardarButton"].click()

       

    except Exception as e:

        log(f"[Advertencia] Fallo al usar diálogo 'Guardar como' (probablemente Win32/UIA): {e}. Usando send_keys como fallback.", "error")

        send_keys(os.path.join(carpeta_export, f"{rname}_export.txt"))

        time.sleep(0.5)

        send_keys('{ENTER}') # O '%g' (Alt+G) si es el atajo de Guardar, pero ENTER es más común

        time.sleep(1)



    # 3. Cerrar msgbox de guardado

    try:

        # Usar la ventana principal 'ventana' para buscar el mensaje

        msg = graph_view.parent().child_window(title_re=".*guardado.*|.*saved.*", control_type="Window")

        if msg.exists(timeout=5):

            log("Msgbox de guardado detectado, cerrando automáticamente...")

            msg.set_focus()

            send_keys('{ENTER}')

            time.sleep(0.2)

    except Exception as e:

        log(f"No se detectó msgbox de guardado o error cerrando: {e}", "debug")



    # 4. Confirmar si falta (si el diálogo pregunta si sobrescribir)

    send_keys('%s') # Alt+S (para "Sí", si aparece una pregunta de sobrescribir)

    log(f"Reporte guardado en: {os.path.join(carpeta_export, f'{rname}_export.txt')}")

    return os.path.join(carpeta_export, f"{rname}_export.txt")
# Asume que esta función log() existe y envía mensajes al dashboard/sistema de logs.
# Ejemplo de placeholder:


def archivo_mas_reciente(carpeta, extension=".txt", rname=None):
    """
    Busca el archivo modificado más recientemente. Si se proporciona rname,
    busca un archivo que contenga ese nombre (ej. 'PAH_export.txt' o 'PAH').
    """
    log(f" Buscando archivos '{extension}' en la carpeta: {carpeta} (Filtro rname: {rname})", "debug")
    
    try:
        archivos_en_carpeta = os.listdir(carpeta)
        
        archivos_completos = [
            os.path.join(carpeta, f) 
            for f in archivos_en_carpeta 
            if f.lower().endswith(extension) and os.path.isfile(os.path.join(carpeta, f))
        ]
        
        archivos_filtrados_por_nombre = archivos_completos
        
        # 2. FILTRADO CRÍTICO: Si existe un nombre base (ej. 'PAH'), buscar archivos que empiecen
        # con él O que contengan el patrón completo (ej. 'PAH_export.txt')
        if rname:
            # CRÍTICO: Usaremos solo el nombre base (ej. 'PAH')
            base_name = rname.split('_export')[0] 
            
            log(f"[debug] Aplicando filtro de nombre base: '{base_name}'", "debug")
            archivos_filtrados_por_nombre = [
                f for f in archivos_completos
                # Busca archivos cuyo nombre empiece con el nombre base
                if os.path.basename(f).startswith(base_name)
            ]
            
            if archivos_filtrados_por_nombre:
                log(f"[debug] Encontrados {len(archivos_filtrados_por_nombre)} archivos que coinciden con '{base_name}'.", "debug")
                archivos_completos = archivos_filtrados_por_nombre
            else:
                log(f" ADVERTENCIA: No se encontró un archivo que coincida con el nombre base '{base_name}'. Buscando el más reciente de TODOS los TXT.", "warn")
        
        log(f"[debug] Archivos elegibles después de filtrar: {len(archivos_completos)}")
        
        if not archivos_completos:
            log(f" No se encontró ningún archivo con la extensión '{extension}' en la carpeta.")
            return None
        
        # 3. Encontrar el más reciente (de la lista de archivos elegibles)
        archivo_reciente = max(archivos_completos, key=os.path.getmtime)
        tiempo_modificacion = os.path.getmtime(archivo_reciente)
        
        log(f" Archivo más reciente identificado: {os.path.basename(archivo_reciente)} (mtime: {tiempo_modificacion})")
        return archivo_reciente
    
    except FileNotFoundError:
        log(f" ERROR: La carpeta '{carpeta}' no fue encontrada.", "error")
        return None
    except Exception as e:
        log(f" ERROR al buscar el archivo más reciente: {e}", "error")
        return None

def esperar_txt_mas_reciente(carpeta, extension=".txt", timeout=60, min_size_bytes=1024, rname=None):
    """
    Espera por un archivo con la extensión dada que sea el más reciente 
    y que supere un tamaño mínimo, dentro de un tiempo de espera (timeout).
    Ahora usa el rname para una búsqueda más precisa.
    """
    log(f" Iniciando espera de archivo '{extension}' en '{carpeta}'. rname='{rname}'. Timeout: {timeout}s, Min Size: {min_size_bytes}B.")
    
    for intento in range(1, timeout + 1):
        log(f" Intento {intento}/{timeout}...")
        
        # Llama a la función actualizada con el nombre base
        archivo = archivo_mas_reciente(carpeta, extension, rname)
        
        if archivo:
            # Archivo encontrado
            tamano_actual = os.path.getsize(archivo)
            log(f"[debug] Archivo encontrado: {os.path.basename(archivo)}. Tamaño actual: {tamano_actual} bytes.")
            
            # Verificación de tamaño mínimo
            if tamano_actual > min_size_bytes:
                log(f" ÉXITO: Archivo encontrado, tamaño suficiente ({tamano_actual} B > {min_size_bytes} B).", "info")
                return archivo
            else:
                log(f" ADVERTENCIA: Tamaño ({tamano_actual} B) insuficiente. Esperando...", "warn")
                
        else:
            # No se encontró archivo en el intento actual
            log(" No se encontró el archivo esperado en el intento actual. Esperando...", "debug")
            
        # Esperar 1 segundo antes del siguiente intento
        time.sleep(1)
        log("[debug] Pausa de 1 segundo completada.")
        
    # Si el bucle termina sin éxito, lanzar la excepción
    log(f" FALLO: Tiempo de espera agotado ({timeout} s). No se encontró el archivo {extension} con el nombre base '{rname}' con el tamaño requerido.", "error")
    raise FileNotFoundError(
        f"No se encontró archivo {extension} en {carpeta} con tamaño > {min_size_bytes} B "
        f"y rname '{rname}' después de {timeout} s."
    )
# -----------------------------
# Flujo principal
# -----------------------------
# ===================================================
# 4. FLUJO PRINCIPAL CORREGIDO
# ===================================================
if __name__ == "__main__":

    log("Logging inicializado correctamente")

    os.environ.setdefault("DJANGO_SETTINGS_MODULE", "muestras.settings")
    django.setup()
    from muestras_crud.models import Empresa, Muestra, Informe 
    from convertir_reportes import txt_a_excel_multi
    from exportar_reporte import excel_a_pdf
    from tests_bet import tests_main
    log("Django inicializado correctamente")
    setup_logging(force=True)
    log("Logging inicializado correctamente DESPUÉS de Django")

    if len(sys.argv) < 4:
        log("=== SCRIPT INICIADO ===")
        log("Uso: python main.py <ruta_exe> <carpeta_qps> <carpeta_export>", "error")
        sys.exit(1)

    PATH_NOVAWIN, CARPETA_QPS, EXPORT_FOLDER = sys.argv[1:4]
    archivos_qps = [os.path.join(r, f) for r, _, fs in os.walk(CARPETA_QPS) for f in fs if f.lower().endswith(".qps")]

    log(f"Encontrados {len(archivos_qps)} archivos para procesar.")

    for i, qps_file in enumerate(archivos_qps, start=1):
        qps_base = os.path.splitext(os.path.basename(qps_file))[0]
        empresa_nombre = qps_base.split("_")[0]
        
        log(f"--- PROCESANDO {i}/{len(archivos_qps)}: {qps_base} ---")
        app = None

        try:
            # 1. Automatización UI
            app, ventana = iniciar_novawin(PATH_NOVAWIN)
            abrir_qps(ventana, qps_file)
            ruta_txt = exportar_reporte(ventana, EXPORT_FOLDER, qps_base)
            
            ruta_txt = esperar_txt_mas_reciente(
              EXPORT_FOLDER,
               rname=qps_base,
                timeout=60,
                  min_size_bytes=1024
                  )
            # 2. Procesamiento de Datos
            log("Procesando archivos y generando Excel/PDF...")
            ruta_excel = txt_a_excel_multi(ruta_txt)
            ruta_excel = renombrar_y_reordenar_hojas(ruta_excel)  # Excel listo

            # 1️Calcular y guardar resultados de tests_main
            resultados_dict = tests_main(ruta_excel, EXPORT_FOLDER)  # Devuelve diccionario
            log(f"main.py[DEBUG] resultados_dict keys: {list(resultados_dict.keys())}", "debug")
            DFT_Poros_norm = normalizar_dft_poros(resultados_dict["DFT_Poros"])

            guardar_resultados_completos(ruta_excel, resultados_dict)  # escribe RESULTADO_CALCULOS
            guardar_resultados_poros(DFT_Poros_norm, ruta_excel, valor="Total")  # escribe RESULTADO_POROS
            output_dir="fractal"
            log("DEBUG: contenido RESULTADO_CALCULOS", "debug")
            df = pd.read_excel(ruta_excel, sheet_name="RESULTADO_CALCULOS")
            log(df.to_string(), "debug")
  
            ruta_frac=fmain(ruta_excel, output_dir="fractal")
            log("despues de modelo fractal")
            log(ruta_frac)

            ruta_pdf = excel_a_pdf(ruta_excel,ruta_frac)
            log("despues de crear el pdf")
            # 3️Subir a Django
            inf = subir_resultados_a_django(
               empresa_nombre=empresa_nombre,
               encargado="Sistema Automático",
               sample_number=qps_base,
               ruta_xlsx=ruta_excel,
               ruta_pdf=ruta_pdf
            )
            log(f"EXITO: Informe ID {inf.id} guardado.", "info")

        except Exception as e:
            log(f"FALLO EN {qps_base}: {str(e)}", "error")
        
        finally:
            cerrar_novawin(app)
            time.sleep(2)

    log("=== PROCESAMIENTO FINALIZADO ===")